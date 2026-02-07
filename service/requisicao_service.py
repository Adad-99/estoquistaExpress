"""
Serviço de criação e gerenciamento de Pedido.
Manipula a lógica de negócio para criar Pedido a partir de planilhas Excel.
"""

import os
import re
from datetime import datetime
from typing import Tuple, Optional
from openpyxl import load_workbook


class PedidoService:
    """Gerencia a criação de Pedido de almoxarifado."""
    
    @staticmethod
    def obter_proximo_numero(pasta: str) -> str:
        """
        Encontra o próximo número de Pedido baseado nos arquivos existentes.
        
        Args:
            pasta: Caminho da pasta para verificar
            
        Returns:
            Próximo número no formato 0001, 0002, etc.
        """
        if not os.path.exists(pasta):
            return "0001"
        
        arquivos = os.listdir(pasta)
        numeros = []
        
        # Procurar por arquivos com padrão de 4 dígitos
        for arquivo in arquivos:
            if arquivo.endswith(('.xlsx', '.xls')):
                # Procurar por padrão de 4 dígitos no nome do arquivo
                match = re.search(r'(\d{4})', arquivo)
                if match:
                    numeros.append(int(match.group(1)))
        
        # Se não encontrou nenhum número, começa com 0001
        if not numeros:
            return "0001"
        
        # Retorna o próximo número
        proximo = max(numeros) + 1
        return f"{proximo:04d}"
    
    @staticmethod
    def criar_Pedido(
        setor: str,
        pasta_destino: str,
        arquivo_padrao: str
    ) -> Tuple[bool, str, Optional[str]]:
        """
        Cria uma nova Pedido copiando e preenchendo a planilha padrão.
        
        Args:
            setor: Nome do setor
            pasta_destino: Pasta onde salvar a Pedido
            arquivo_padrao: Caminho da planilha padrão
            
        Returns:
            Tupla (sucesso, mensagem, caminho_arquivo)
        """
        try:
            # Validações
            if not setor or not setor.strip():
                return False, "Setor não informado", None
            
            if not os.path.exists(pasta_destino):
                return False, f"Pasta de destino não encontrada: {pasta_destino}", None
            
            if not os.path.exists(arquivo_padrao):
                return False, f"Arquivo padrão não encontrado: {arquivo_padrao}", None
            
            # Obter próximo número
            numero_Pedido = PedidoService.obter_proximo_numero(pasta_destino)
            
            # Carregar arquivo padrão
            wb = load_workbook(arquivo_padrao)
            ws = wb.active
            
            # Preencher dados
            ws['C4'] = setor.strip()  # Setor
            ws['H4'] = numero_Pedido  # Número da Pedido
            ws['B6'] = datetime.now().strftime('%d/%m/%Y')  # Data atual
            
            # Salvar arquivo
            nome_arquivo = f"{numero_Pedido}.xlsx"
            caminho_completo = os.path.join(pasta_destino, nome_arquivo)
            wb.save(caminho_completo)
            
            mensagem = (
                f"Pedido criada com sucesso!\n\n"
                f"Número: {numero_Pedido}\n"
                f"Setor: {setor}\n"
                f"Arquivo: {nome_arquivo}"
            )
            
            return True, mensagem, caminho_completo
            
        except Exception as e:
            return False, f"Erro ao criar Pedido: {str(e)}", None
    
    @staticmethod
    def validar_planilha_padrao(arquivo: str) -> Tuple[bool, str]:
        """
        Valida se o arquivo pode ser usado como planilha padrão.
        
        Args:
            arquivo: Caminho do arquivo a validar
            
        Returns:
            Tupla (válido, mensagem)
        """
        if not os.path.exists(arquivo):
            return False, "Arquivo não encontrado"
        
        if not arquivo.lower().endswith(('.xlsx', '.xls')):
            return False, "Arquivo deve ser do tipo Excel (.xlsx ou .xls)"
        
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
            
            # Verificar se possui as células necessárias
            # Não é necessário que estejam preenchidas, apenas que existam
            if ws is None:
                return False, "Planilha não possui uma aba ativa"
            
            return True, "Planilha válida"
            
        except Exception as e:
            return False, f"Erro ao validar planilha: {str(e)}"
